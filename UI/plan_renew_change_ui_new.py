import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import UI.shared_data as shared_data
import pages.login_page as test_login


def fetch_client_details(client_name, instance, file_path=r"C:path\data\Client_data.xlsx"):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        header = [cell.value for cell in sheet[1]]
        instance_url_col = f"{instance} URL"
        instance_password_col = f"Password {instance}"

        if instance_url_col not in header or instance_password_col not in header or "Username" not in header:
            print("Required columns are missing.")
            return None, None, None

        url_col_index = header.index(instance_url_col)
        username_col_index = header.index("Username")
        password_col_index = header.index(instance_password_col)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0].strip().lower() == client_name.strip().lower():
                url = row[url_col_index]
                username = row[username_col_index]
                password = row[password_col_index]
                shared_data.url = url
                shared_data.password = password
                shared_data.username = username
                return url, username, password

        return None, None, None

    except Exception as e:
        print(f"Error fetching client details: {e}")
        return None, None, None


def load_client_names(file_path=r"C:pathautomation_framework\data\Client_data.xlsx"):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        clients = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            client_name = row[0]
            if client_name:
                clients.add(client_name.strip())
        return sorted(list(clients))
    except Exception as e:
        print(f"Error loading clients: {e}")
        return []


def get_url_for_client_and_instance(client_name, instance, file_path=r"C:path\automation_framework\data\Client_data.xlsx"):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        header = [cell.value for cell in sheet[1]]
        instance_col_header = instance + " URL"
        if instance_col_header not in header:
            return None
        instance_col_index = header.index(instance_col_header)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0].strip().lower() == client_name.strip().lower():
                return row[instance_col_index]
        return None
    except Exception as e:
        print(f"Error reading URL from Excel: {e}")
        return None


# Plan renew change Functions
def run_plan_change(client, instance, customer_id, url):
    shared_data.client = client
    shared_data.instance = instance
    shared_data.customer_id = customer_id
    shared_data.url = url
    return f"[{client} | {instance} | {url}] Plan change test run for {customer_id}"


def run_plan_renew(client, instance, customer_id, url, username, password):
    shared_data.client = client
    shared_data.instance = instance
    shared_data.customer_id = customer_id
    shared_data.username = username
    shared_data.password = password
    if url and not url.startswith("http"):
        url = "https://" + url
    shared_data.url = url

    print(f"[DEBUG] Running plan renew with URL: {url}")
    test_login.main()

    return f"[{client} | {instance} | {url}] Plan renew test run for {customer_id}"


def start_ui():
    root = tk.Tk()
    root.title("Automation Framework UI")
    root.geometry("520x600")
    root.resizable(False, False)
    root.configure(bg="#f4f4f4")

    main_frame = tk.Frame(root, bg="#f4f4f4", padx=20, pady=20)
    main_frame.pack(expand=True, fill="both")

    # --- Client Dropdown ---
    tk.Label(main_frame, text="Select Client", font=("Segoe UI", 11, "bold"),
             bg="#f4f4f4", fg="#333").grid(row=0, column=0, columnspan=2, sticky="w")

    client_names = load_client_names()
    client_dropdown = ttk.Combobox(main_frame, values=client_names, state="readonly", width=40)
    client_dropdown.set("Select Client")
    client_dropdown.grid(row=1, column=0, columnspan=2, pady=5, sticky="ew")

    # --- Instance Dropdown ---
    tk.Label(main_frame, text="Select Instance", font=("Segoe UI", 11, "bold"),
             bg="#f4f4f4", fg="#333").grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))

    instance_options = ["QA", "Demo", "Live"]
    instance_dropdown = ttk.Combobox(main_frame, values=instance_options, state="readonly", width=40)
    instance_dropdown.set("Select Instance")
    instance_dropdown.grid(row=3, column=0, columnspan=2, pady=5, sticky="ew")

    # --- Customer ID Entry & Radio Buttons ---
    tk.Label(main_frame, text="Customer Selection Options", font=("Segoe UI", 11, "bold"),
             bg="#f4f4f4", fg="#333").grid(row=4, column=0, columnspan=2, sticky="w", pady=(10, 5))

    customer_option = tk.StringVar(value="manual")
    placeholder_text = "Enter Customer ID"

    customer_id_entry = tk.Entry(main_frame, width=40, fg='grey')
    customer_id_entry.insert(0, placeholder_text)
    customer_id_entry.grid(row=5, column=0, columnspan=2, pady=(0, 10), sticky="ew")

    def clear_placeholder(event):
        if customer_id_entry.get() == placeholder_text:
            customer_id_entry.delete(0, tk.END)
            customer_id_entry.config(fg='black')

    def set_placeholder(event):
        if customer_id_entry.get() == "":
            customer_id_entry.insert(0, placeholder_text)
            customer_id_entry.config(fg='grey')

    customer_id_entry.bind("<FocusIn>", clear_placeholder)
    customer_id_entry.bind("<FocusOut>", set_placeholder)

    def update_customer_entry_state():
        is_manual = customer_option.get() == "manual"
        state = "normal" if is_manual else "disabled"
        customer_id_entry.config(state=state)
        if not is_manual:
            customer_id_entry.delete(0, tk.END)
            customer_id_entry.insert(0, placeholder_text)
            customer_id_entry.config(fg='grey')

    radio_options = [
        ("Enter customer ID", "manual"),
        ("Search for an active customer from TG5", "active_one"),
        ("Search for all active customers from TG5", "active_all"),
        ("Search for a suspended customer from TG5", "suspended_one"),
        ("Search for all suspended customers from TG5", "suspended_all"),
    ]

    for i, (label, value) in enumerate(radio_options):
        rb = tk.Radiobutton(
            main_frame, text=label, variable=customer_option, value=value,
            bg="#f4f4f4", anchor="w", font=("Segoe UI", 9),
            command=update_customer_entry_state
        )
        rb.grid(row=6 + i, column=0, columnspan=2, sticky="w", pady=2)

    # --- Buttons ---
    def handle_run(action_type):
        client = client_dropdown.get()
        instance = instance_dropdown.get()
        option = customer_option.get()
        customer_id = customer_id_entry.get().strip()

        if client == "Select Client" or instance == "Select Instance":
            messagebox.showerror("Error", "Please select both client and instance.")
            return

        if option == "manual":
            if not customer_id or customer_id == placeholder_text:
                messagebox.showerror("Error", "Please enter a valid Customer ID.")
                return
            shared_data.customer_option = "manual"
            shared_data.customer_id = customer_id
        else:
            shared_data.customer_option = option
            shared_data.customer_id = ""

        if action_type == "renew":
            url, username, password = fetch_client_details(client, instance)
            if not url or not username or not password:
                messagebox.showerror("Error", "Client credentials or URL not found.")
                return
            result = run_plan_renew(client, instance, shared_data.customer_id, url, username, password)
        else:
            url = get_url_for_client_and_instance(client, instance)
            if not url:
                messagebox.showerror("Error", "URL not found for the selected client and instance.")
                return
            result = run_plan_change(client, instance, shared_data.customer_id, url)

        print("-----------------------------------------------------------------------------------")
        print(f"Customer Option: {shared_data.customer_option}")
        print(f"Customer ID: {shared_data.customer_id}")

        messagebox.showinfo("Success", result)

    button_row = 6 + len(radio_options)

    btn_change = tk.Button(main_frame, text="Run Plan Change", font=("Segoe UI", 10, "bold"),
                           bg="#007acc", fg="white", command=lambda: handle_run("change"))
    btn_change.grid(row=button_row, column=0, pady=(20, 10), sticky="ew")

    btn_renew = tk.Button(main_frame, text="Run Plan Renew", font=("Segoe UI", 10, "bold"),
                          bg="#28a745", fg="white", command=lambda: handle_run("renew"))
    btn_renew.grid(row=button_row, column=1, pady=(20, 10), sticky="ew")

    root.mainloop()


if __name__ == "__main__":
    start_ui()
