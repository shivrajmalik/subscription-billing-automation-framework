from flask import Flask, render_template, request, redirect, flash
import openpyxl
import os
import UI.shared_data as shared_data
import pages.login_page as test_login

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Required for flashing messages

EXCEL_PATH = r"C:path\automation_framework\data\Client_data.xlsx"

def load_client_names(file_path=EXCEL_PATH):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        clients = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            client_name = row[0]
            if client_name:
                clients.add(client_name.strip())
        return sorted(list(clients))
    except Exception as e:
        print(f"Error loading clients: {e}")
        return []

def get_url_for_client_and_instance(client_name, instance, file_path=EXCEL_PATH):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        header = [cell.value for cell in sheet[1]]
        instance_col_header = instance + " URL"
        if instance_col_header not in header:
            return None
        instance_col_index = header.index(instance_col_header)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0].strip().lower() == client_name.strip().lower():
                return row[instance_col_index]
        return None
    except Exception as e:
        print(f"Error reading URL from Excel: {e}")
        return None

def run_plan_change(client, instance, customer_id, url):
    shared_data.client = client
    shared_data.instance = instance
    shared_data.customer_id = customer_id
    shared_data.url = url
    return f"[{client} | {instance} | {url}] Plan change test run for {customer_id}"

def run_plan_renew(client, instance, customer_id, url):
    shared_data.client = client
    shared_data.instance = instance
    shared_data.customer_id = customer_id
    if url and not url.startswith("http"):
        url = "https://" + url
    shared_data.url = url

    print(f"[DEBUG] Running plan renew with URL: {url}")
    test_login.main()

    return f"[{client} | {instance} | {url}] Plan renew test run for {customer_id}"

@app.route("/", methods=["GET", "POST"])
def index():
    client_names = load_client_names()
    selected_client = request.form.get("client")
    selected_instance = request.form.get("instance")
    selected_option = request.form.get("customer_option", "manual")
    customer_id = request.form.get("customer_id", "").strip()
    action_type = request.form.get("action_type")

    if request.method == "POST":
        if selected_client == "Select Client" or selected_instance == "Select Instance":
            flash("Please select both client and instance.", "danger")
        elif selected_option == "manual" and (not customer_id or customer_id.lower() == "enter customer id"):
            flash("Please enter a valid customer ID.", "danger")
        else:
            url = get_url_for_client_and_instance(selected_client, selected_instance)
            if not url:
                flash("URL not found for selected client and instance.", "danger")
            else:
                shared_data.client = selected_client
                shared_data.instance = selected_instance
                shared_data.customer_option = selected_option
                shared_data.customer_id = customer_id if selected_option == "manual" else ""
                shared_data.url = url

                print("-----------------------------------------------------------------------------------")
                print(f"Customer Option: {shared_data.customer_option}")
                print(f"Customer ID: {shared_data.customer_id}")

                if action_type == "change":
                    result = run_plan_change(selected_client, selected_instance, shared_data.customer_id, url)
                else:
                    result = run_plan_renew(selected_client, selected_instance, shared_data.customer_id, url)

                flash(result, "success")
                return redirect("/")  # Avoid form re-submission on refresh

    return render_template("index.html",
                           client_names=client_names,
                           selected_client=selected_client,
                           selected_instance=selected_instance,
                           selected_option=selected_option,
                           customer_id=customer_id)

if __name__ == "__main__":
    app.run(debug=True)
